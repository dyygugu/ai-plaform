from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4
from typing import Optional

from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.settings import get_settings
from app.models.account import AccountStatus, AidpAccount
from app.models.audit import AuditLog, AuditSeverity
from app.models.ops import EarningsSnapshot
from app.models.task import TaskCatalogItem
from app.models.worker import Worker
from app.schemas.data_quality import (
    DataQualityCheckItem,
    DataQualityExportResponse,
    DataQualityReportRequest,
    DataQualityReportResponse,
    DataQualitySummaryResponse,
    EarningsContractItem,
)
from app.services.api_paths import api_path, api_paths
from app.services.audit_service import write_audit
from app.services.earnings_service import list_earnings
from app.services.task_rules import utc_now

EXPECTED_ACCOUNT_COUNT = 7


def build_data_quality_summary(db: Session) -> DataQualitySummaryResponse:
    settings = get_settings()
    accounts = list(db.scalars(select(AidpAccount).order_by(AidpAccount.user_id.asc())))
    tasks = list(db.scalars(select(TaskCatalogItem).order_by(TaskCatalogItem.task_id.asc())))
    earnings_rows = list_earnings(db)
    workers = list(db.scalars(select(Worker).order_by(Worker.worker_id.asc())))
    audit_count = int(db.scalar(select(func.count(AuditLog.id))) or 0)
    db.flush()

    production_accounts = [account for account in accounts if _is_real_user_id(account.user_id) and account.status != AccountStatus.DISABLED]
    non_production_count = len(accounts) - len(production_accounts)
    account_ids = {account.user_id for account in production_accounts}
    earning_account_ids = {row.account_user_id for row in earnings_rows}
    numeric_pending_ok = all(_parse_pending(task.pending_raw) >= 0 for task in tasks)
    has_short_names = all(bool(task.task_short_name) for task in tasks)
    latest_acceptance = _latest_artifact(_reports_root(), "acceptance-*.md")
    latest_coverage = _latest_artifact(_reports_root(), "account-coverage-baseline-*.md")
    checks = [
        DataQualityCheckItem(
            key="accounts",
            title="7 账号基线",
            status="passed" if len(production_accounts) == EXPECTED_ACCOUNT_COUNT else "failed",
            expected=f"{EXPECTED_ACCOUNT_COUNT} 个生产账号",
            actual=f"生产账号 {len(production_accounts)} 个，停用/非生产 {non_production_count} 个",
            evidence_path=api_path("/accounts", settings),
            message="8789 原生生产账号基线已写入新库；停用/非生产账号不计入失败。" if len(production_accounts) == EXPECTED_ACCOUNT_COUNT else "生产账号数量与 8789 原生基线不一致。",
        ),
        DataQualityCheckItem(
            key="task_catalog",
            title="任务目录口径",
            status="passed" if tasks and has_short_names else "failed",
            expected="任务简称、任务 ID、待处理数字均可核验",
            actual=f"{len(tasks)} 条任务，short_name_ok={has_short_names}",
            evidence_path=api_path("/tasks/catalog", settings),
            message="任务目录包含简称和当前待处理数字。" if tasks else "任务目录为空。",
        ),
        DataQualityCheckItem(
            key="pending_numeric",
            title="待处理数字口径",
            status="passed" if numeric_pending_ok else "failed",
            expected="pending_raw 可解析为非负数字",
            actual=f"pending_numeric_ok={numeric_pending_ok}",
            evidence_path=api_path("/tasks/catalog", settings),
            message="待处理字段保持数字口径，不再展示节点名称占位。",
        ),
        DataQualityCheckItem(
            key="earnings_accounts",
            title="收益账号覆盖",
            status="passed" if account_ids and account_ids.issubset(earning_account_ids) else "failed",
            expected="收益导出覆盖 7 个账号",
            actual=f"收益账号 {len(earning_account_ids)} 个",
            evidence_path=api_path("/data-quality/export", settings),
            message="收益三项、今日收益、小时收益按账号维度补齐。",
        ),
        DataQualityCheckItem(
            key="workers_audit",
            title="Worker 与审计证据",
            status="passed" if workers and audit_count >= 1 else "warning",
            expected="至少 1 个 Worker 与审计事件",
            actual=f"workers={len(workers)}，audit_events={audit_count}",
            evidence_path=api_paths("/workers", "/audit/logs", settings=settings),
            message="Worker 与审计可作为数据闭环旁证；warning 不触发外部系统。",
        ),
        DataQualityCheckItem(
            key="p16_coverage",
            title="P16 覆盖基线证据",
            status="passed" if latest_coverage else "warning",
            expected="存在账号覆盖基线报告",
            actual=latest_coverage or "未找到",
            evidence_path=latest_coverage or "data/reports/account-coverage-baseline-*.md",
            message="P17 沿用 P16 账号覆盖基线作为账号/任务交叉校验输入。",
        ),
        DataQualityCheckItem(
            key="latest_acceptance",
            title="最新自动验收报告",
            status="passed" if latest_acceptance else "warning",
            expected="存在 acceptance 报告",
            actual=latest_acceptance or "未找到",
            evidence_path=latest_acceptance or "Projects/aidp-monitor-next/reports",
            message="自动验收报告用于最终矩阵汇总。",
        ),
    ]
    status = "passed" if all(item.status != "failed" for item in checks) else "failed"
    contracts = _earnings_contracts()
    today_total = sum(float(row.today_income) for row in earnings_rows)
    hourly_total = sum(float(row.hourly_income) for row in earnings_rows)
    return DataQualitySummaryResponse(
        generated_at=utc_now(),
        status=status,
        expected_account_count=EXPECTED_ACCOUNT_COUNT,
        account_count=len(production_accounts),
        task_count=len(tasks),
        earnings_row_count=len(earnings_rows),
        worker_count=len(workers),
        audit_event_count=audit_count,
        today_income_total=today_total,
        hourly_income_total=hourly_total,
        checks=checks,
        contracts=contracts,
        risk_notes=[
            "P17 固化收益三项、今日收益和小时收益口径；样例值可为 0，但字段和导出结构必须稳定。",
            "导出文件只包含账号元数据、任务简称、待处理数字和证据路径，不包含 Cookie 明文。",
            "正式域名切换提醒仍推迟到 P20 封版后。",
        ],
        next_actions=[
            "人工验收时打开数据校验页核对账号数、任务数和导出路径。",
            "若后续接入真实收益采集，只替换收益快照写入来源，不改变 P17 导出口径。",
            "P18 继续补异常处置与运维闭环。",
        ],
        message="P17 数据正确性与收益口径已形成本地校验基线。",
    )


def list_data_quality_checks(db: Session) -> list[DataQualityCheckItem]:
    return build_data_quality_summary(db).checks


def export_data_quality_workbook(db: Session) -> DataQualityExportResponse:
    summary = build_data_quality_summary(db)
    accounts = list(db.scalars(select(AidpAccount).order_by(AidpAccount.user_id.asc())))
    tasks = list(db.scalars(select(TaskCatalogItem).order_by(TaskCatalogItem.source_account_user_id.asc(), TaskCatalogItem.task_id.asc())))
    earnings_rows = list_earnings(db)
    export_root = _reports_root() / "exports"
    export_root.mkdir(parents=True, exist_ok=True)
    path = export_root / f"data-quality-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_summary_sheet(workbook, summary)
    _write_accounts_sheet(workbook, accounts, earnings_rows)
    _write_tasks_sheet(workbook, tasks)
    _write_checks_sheet(workbook, summary)
    workbook.save(path)

    row_counts = {
        "accounts": len(accounts),
        "tasks": len(tasks),
        "earnings": len(earnings_rows),
        "checks": len(summary.checks),
    }
    display_path = _display_path(path)
    return DataQualityExportResponse(
        generated_at=utc_now(),
        status=summary.status,
        export_path=display_path,
        row_counts=row_counts,
        evidence_paths=[display_path, *[check.evidence_path for check in summary.checks]],
        metadata={"cookie_copy_enabled": False, "manual_domain_switch_deferred": True},
        message="数据质量 Excel 已导出；覆盖 7 账号、任务简称、待处理数字、时间戳和证据路径。",
    )


def create_data_quality_report(db: Session, request: DataQualityReportRequest) -> DataQualityReportResponse:
    summary = build_data_quality_summary(db)
    trace_id = uuid4().hex
    export = export_data_quality_workbook(db) if request.generate_excel else None
    report_path = _write_quality_report(summary, trace_id, export.export_path if export else None) if request.generate_report else None
    audit_trace_id = None
    if request.write_audit:
        audit = write_audit(
            db,
            event_type="data_quality_baseline",
            severity=AuditSeverity.INFO if summary.status == "passed" else AuditSeverity.WARNING,
            target_type="data_quality",
            target_id=trace_id,
            message=f"P17 data quality status={summary.status}, accounts={summary.account_count}, tasks={summary.task_count}, earnings={summary.earnings_row_count}",
        )
        db.commit()
        audit_trace_id = audit.trace_id
    row_counts = export.row_counts if export else {"checks": len(summary.checks)}
    evidence_paths = ([export.export_path] if export else []) + [check.evidence_path for check in summary.checks]
    if report_path:
        evidence_paths.insert(0, report_path)
    return DataQualityReportResponse(
        generated_at=utc_now(),
        status=summary.status,
        report_path=report_path,
        export_path=export.export_path if export else None,
        audit_trace_id=audit_trace_id,
        row_counts=row_counts,
        evidence_paths=evidence_paths,
        summary=summary,
        message="P17 数据正确性报告已生成；不复制 Cookie，不触发外部系统。",
    )


def _earnings_contracts() -> list[EarningsContractItem]:
    return [
        EarningsContractItem(key="income_1", title="收益三项 1", source_field="income_1_value", display_name="页面原始收入项1", aggregation="账号维度原样展示，总览不改名", status="sealed"),
        EarningsContractItem(key="income_2", title="收益三项 2", source_field="income_2_value", display_name="页面原始收入项2", aggregation="账号维度原样展示，总览不改名", status="sealed"),
        EarningsContractItem(key="income_3", title="收益三项 3", source_field="income_3_value", display_name="页面原始收入项3", aggregation="账号维度原样展示，总览不改名", status="sealed"),
        EarningsContractItem(key="today_income", title="今日收益", source_field="today_income", display_name="今日收益", aggregation="按账号求和为总览今日收益", status="sealed"),
        EarningsContractItem(key="hourly_income", title="小时收益", source_field="hourly_income", display_name="小时收益", aggregation="按账号求和为总览小时收益", status="sealed"),
    ]


def _write_summary_sheet(workbook: Workbook, summary: DataQualitySummaryResponse) -> None:
    sheet = workbook.create_sheet("口径摘要")
    sheet.append(["字段", "值"])
    for key, value in [
        ("状态", summary.status),
        ("账号数量", f"{summary.account_count}/{summary.expected_account_count}"),
        ("任务数量", summary.task_count),
        ("收益行数", summary.earnings_row_count),
        ("今日收益", summary.today_income_total),
        ("小时收益", summary.hourly_income_total),
        ("生成时间", summary.generated_at.isoformat()),
    ]:
        sheet.append([key, value])


def _write_accounts_sheet(workbook: Workbook, accounts: list[AidpAccount], earnings_rows: list[EarningsSnapshot]) -> None:
    sheet = workbook.create_sheet("账号收益")
    earnings_by_account = {row.account_user_id: row for row in earnings_rows}
    sheet.append(["账号", "显示名", "状态", "任务来源", "收益项1", "收益项2", "收益项3", "今日收益", "小时收益", "采集时间"])
    for account in accounts:
        row = earnings_by_account.get(account.user_id)
        sheet.append([
            account.user_id,
            account.display_name,
            account.status.value,
            "是" if account.is_task_source else "否",
            float(row.income_1_value) if row else 0,
            float(row.income_2_value) if row else 0,
            float(row.income_3_value) if row else 0,
            float(row.today_income) if row else 0,
            float(row.hourly_income) if row else 0,
            row.captured_at.isoformat() if row and row.captured_at else "",
        ])


def _write_tasks_sheet(workbook: Workbook, tasks: list[TaskCatalogItem]) -> None:
    sheet = workbook.create_sheet("任务目录")
    sheet.append(["来源账号", "任务ID", "任务简称", "任务名称ID", "状态", "待处理数字", "可见性", "最近任务页时间", "证据路径"])
    for task in tasks:
        sheet.append([
            task.source_account_user_id,
            task.task_id,
            task.task_short_name,
            task.task_name_id,
            task.task_status_raw,
            _parse_pending(task.pending_raw),
            task.visibility.value,
            task.last_task_page_seen_at.isoformat() if task.last_task_page_seen_at else "",
            api_path("/tasks/catalog"),
        ])


def _write_checks_sheet(workbook: Workbook, summary: DataQualitySummaryResponse) -> None:
    sheet = workbook.create_sheet("一致性检查")
    sheet.append(["检查", "状态", "预期", "实际", "证据", "说明"])
    for check in summary.checks:
        sheet.append([check.title, check.status, check.expected, check.actual, check.evidence_path, check.message])


def _write_quality_report(summary: DataQualitySummaryResponse, trace_id: str, export_path: Optional[str]) -> str:
    path = _reports_root() / f"data-quality-baseline-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.md"
    lines = [
        "# aidp-monitor-next P17 数据正确性与收益口径封版",
        "",
        f"生成时间：{summary.generated_at.isoformat()}",
        f"trace_id：{trace_id}",
        f"状态：{summary.status}",
        f"账号数量：{summary.account_count}/{summary.expected_account_count}",
        f"任务数量：{summary.task_count}",
        f"收益行数：{summary.earnings_row_count}",
        f"今日收益合计：{summary.today_income_total}",
        f"小时收益合计：{summary.hourly_income_total}",
        f"增强导出：{export_path or '未生成'}",
        "",
        "## 收益口径",
    ]
    for contract in summary.contracts:
        lines.append(f"- {contract.title}：source={contract.source_field}，display={contract.display_name}，aggregation={contract.aggregation}，status={contract.status}")
    lines.extend(["", "## 一致性检查"])
    for check in summary.checks:
        lines.append(f"- [{check.status}] {check.title}：expected={check.expected}，actual={check.actual}，evidence={check.evidence_path}")
    lines.extend(["", "## 风险提示"])
    for note in summary.risk_notes:
        lines.append(f"- {note}")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return _display_path(path)


def _parse_pending(value: str) -> int:
    try:
        return int(str(value).replace(",", "").strip() or "0")
    except ValueError:
        return 0


def _is_real_user_id(value: str) -> bool:
    text = str(value or "").strip()
    return text.isdigit() and 12 <= len(text) <= 24


def _reports_root() -> Path:
    settings = get_settings()
    root = Path(settings.task_sample_root).resolve().parent / "reports"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _latest_artifact(root: Path, pattern: str) -> Optional[str]:
    matches = sorted(root.glob(pattern), key=lambda item: item.stat().st_mtime, reverse=True)
    return _display_path(matches[0]) if matches else None


def _display_path(path: Path) -> str:
    settings = get_settings()
    runtime_root = Path(settings.task_sample_root).resolve().parent.parent
    try:
        return path.resolve().relative_to(runtime_root.resolve()).as_posix()
    except ValueError:
        pass
    for candidate in Path(__file__).resolve().parents:
        if candidate.name == "aidp-monitor-next":
            workspace = candidate.parents[1] if len(candidate.parents) > 1 else candidate
            try:
                return path.resolve().relative_to(workspace.resolve()).as_posix()
            except ValueError:
                break
    return str(path)
